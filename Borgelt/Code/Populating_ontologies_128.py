import emmopy
import Cython
import collections
import os
import types
import json
import collections
from collections import OrderedDict
from ontopy import World
from ontopy.utils import write_catalog
import owlready2
import pandas as pd
import numpy as np
import time
from mergedeep import merge
import matplotlib.pyplot as plt
# import plotly.express as px
# import plotly.graph_objects as go
import kaleido
import pymongo
from pymongo import MongoClient
from Test_for_adding_classes_to_ontology_functions_2 import *
import translator_list_json_to_onto_2
import copy
from openpyxl import load_workbook

# mongodb+srv://HendrikB:tAyJ0AtmHUiqaYNf@cluster0.4a1ca.mongodb.net/test
# cluster = MongoClient('mongodb+srv://HendrikB:tAyJ0AtmHUiqaYNf@cluster0.4a1ca.mongodb.net/myFirstDatabase?retryWrites=true&w=majority')
# db = cluster['MasterthesisDB']
# mongo_collection = db['JSONdicts']

t_2 = time.time()

def list_of_emmo_classes():
    list_emmo_classes = list(emmo.classes())
    for i_1, i_2 in enumerate(list_emmo_classes):
        list_emmo_classes[i_1] = ''.join(i_2.prefLabel)
        list_emmo_classes[i_1] = [str(list_emmo_classes[i_1]), i_2]
    return list_emmo_classes

f_3 = OrderedDict()
for name in os.listdir('C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis'):
    if name.endswith("Dict_of_all_keys.json"):
        with open(name) as outfile:
            f_3.update({name.rstrip('.json'): json.load(outfile)})

Dict_of_all_key = f_3['Dict_of_all_key']

Dict_of_Domain = Dict_of_all_key['CFX Command Language for Run']['Initialization']['FLOW']['Flow Analysis 1']['DOMAIN']

world = World()

owlready2.JAVA_EXE = 'C:/Users/hendr/Desktop/Protege-5.5.0/jre/bin/java.exe'

owlready2.onto_path.append('C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis')
emmo = world.get_ontology('emmo_inferred_20220216_trying_classes_2.rdf')
emmo.name = 'emmo'
emmo.load()
emmo.sync_python_names()
emmo.base_iri = "http://emmo.info/emmo#"
namespace = emmo.get_namespace("http://emmo.info/emmo#")



# # here is the basic way to create classes by starting with initializing onto, then defining a variable for further steps and then using the function types.new_classes('Class_Name', (BaseClasses/SuperClasses, ...))
# with emmo:
#     namespace = 'emmo'
#     # classes
#     Domains = types.new_class('Domains', (emmo.Symbolic,))
#     emmo.Domains.prefLabel = 'Domains'
#     # ProgramSettings = types.new_class('ProgramSettings', (emmo.Symbolic,))
#     # emmo.ProgramSettings.prefLabel = 'ProgramSettings'
#     BoundarySettings = types.new_class('BoundarySettings', (emmo.Symbolic,))
#     emmo.BoundarySettings.prefLabel = 'BoundarySettings'
#     Boundaries = types.new_class('Boundaries', (emmo.BoundarySettings,))
#     emmo.Boundaries.prefLabel = 'Boundaries'
#     Inlet = types.new_class('Inlet', (emmo.Boundaries,))
#     emmo.Inlet.prefLabel = 'Inlet'
#     Outlet = types.new_class('Outlet', (emmo.Boundaries,))
#     emmo.Outlet.prefLabel = 'Outlet'
#     Wall = types.new_class('Wall', (emmo.Boundaries,))
#     emmo.Wall.prefLabel = 'Wall'
#     # Interface = types.new_class('Interface', (emmo.Boundaries,))
#     # BoundaryConditions = types.new_class('BoundaryConditions', (emmo.BoundarySettings,))
#     # BoundaryConditionsComponent = types.new_class('BoundaryConditionsComponent', (emmo.BoundaryConditions,))
#     # BoundaryConditionsFlow = types.new_class('BoundaryConditionsFlow', (emmo.BoundaryConditions,))
#     # BoundaryConditionsMassAndMomentum = types.new_class('BoundaryConditionsFlow', (emmo.BoundaryConditions,))
#
#     # Object properties
#     hasBoundarySettings = types.new_class('hasBoundarySettings', (owlready2.ObjectProperty,))
#     hasBoundarySettings.range = [emmo.Domains]
#     hasBoundarySettings.domain = [emmo.BoundarySettings]
#
#     hasBoundaries = types.new_class('hasBoundaries', (emmo.hasBoundarySettings,))
#
#     InletBoundaryIndividual = Inlet('InletBoundaryIndividual')
#     OutletBoundaryIndividual = Outlet('OutletBoundaryIndividual')
#     WallBoundaryIndividual = Wall('WallBoundaryIndividual')
#
# list_of_emmo_individuals(emmo)
#
# BoundarySettingIndividual = emmo.BoundarySettings('BoundarySettingIndividual', hasBoundaries=[InletBoundaryIndividual, OutletBoundaryIndividual])
# # BoundarySettingIndividual_2 = BoundarySettings(hasBoundaries=[WallBoundaryIndividual])
# a_123 = list(BoundarySettingIndividual.hasBoundaries)
# a_234 = search_for_index('InletBoundaryIndividual', list_of_emmo_individuals(emmo))
# a_345 = search_for_index('OutletBoundaryIndividual', list_of_emmo_individuals(emmo))
#
# a_456 = search_for_index('OutletBoundaryIndividual', list_of_emmo_individuals(emmo))
# a_567 = a_123.append(a_456)
# a_678 = list_of_emmo_individuals(emmo)
# emmo.BoundarySettingIndividual.hasBoundaries.append(WallBoundaryIndividual)
# emmo.BoundarySettingIndividual.is_a.append(hasBoundaries.value(WallBoundaryIndividual))

# BoundarySettingIndividual.is_a(hasBoundaries.has_self(WallBoundaryIndividual))
#
# test_list_of_emmo_classes = list_of_emmo_classes()
# emmo.save(os.path.join('C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis', 'emmo_inferred_20220216_trying_classes.owl'))

# list_of_translator_keys = append_key_to_list(f_3['Dict_of_all_key']['CFX Command Language for Run']['Initialization']['FLOW']['Flow Analysis 1']['DOMAIN'])
# list_of_translator_keys = append_key_to_list(f_3['Dict_of_all_key']['CFX Command Language for Run'])
# pd.DataFrame(list_of_translator_keys).to_excel('output_2.xlsx', header=False, index=False)


list_of_keys_containing_named_instances = translator_list_json_to_onto_2.list_of_keys_containing_named_instances
list_of_keys_containing_named_sub_instances = translator_list_json_to_onto_2.list_of_keys_containing_named_sub_instances
list_of_keys_containing_named_sub_sub_instances = translator_list_json_to_onto_2.list_of_keys_containing_named_sub_sub_instances
list_of_keys_containing_named_sub_sub_sub_instances = translator_list_json_to_onto_2.list_of_keys_containing_named_sub_sub_sub_instances
list_of_instance_keys = translator_list_json_to_onto_2.list_of_instance_keys
list_of_instance_sub_keys = translator_list_json_to_onto_2.list_of_instance_sub_keys
list_of_instance_sub_sub_keys = translator_list_json_to_onto_2.list_of_instance_sub_sub_keys
# list_of_all_instance_keys =
all_sims = OrderedDict()
for name in os.listdir('C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/Json_Dict_unmodified'):
    if name.endswith(".json"):
        with open('./Json_Dict_unmodified/' + name) as outfile:
            all_sims.update({name.rstrip('.json'): json.load(outfile)})

dict_of_key_sections = {'list_of_individual_sub_sub_keys': list_of_instance_sub_sub_keys,
                        'list_of_individual_sub_keys': list_of_instance_sub_keys,
                        'list_of_individual_keys': list_of_instance_keys}

t = time.time()
modified_all_sims = copy.deepcopy(all_sims)
search_for_name_source_in_dict(modified_all_sims)

save_and_archiv_named_sub_sub_sub_instances(modified_all_sims, list_of_keys_containing_named_sub_sub_sub_instances)
sub_sub_sub_dict_folder_path = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/dict_inst_2/Named_sub_sub_sub_inst_2/'
new_dict_all_sims_sub_sub_sub = change_name_of_sub_sub_inst(modified_all_sims, sub_sub_sub_dict_folder_path)
save_and_archiv_named_sub_sub_instances(new_dict_all_sims_sub_sub_sub, list_of_keys_containing_named_sub_sub_instances)
sub_sub_dict_folder_path = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/dict_inst_2/Named_sub_sub_inst_2/'
new_dict_all_sims_sub_sub = change_name_of_sub_sub_inst(new_dict_all_sims_sub_sub_sub, sub_sub_dict_folder_path)
save_and_archiv_named_sub_instances_2(new_dict_all_sims_sub_sub, list_of_keys_containing_named_sub_instances)
sub_dict_folder_path = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/dict_inst_2/Named_sub_inst_2/'
new_dict_all_sims_sub = change_name_of_sub_sub_inst(new_dict_all_sims_sub_sub, sub_dict_folder_path)
save_and_archiv_named_instances(new_dict_all_sims_sub, list_of_keys_containing_named_instances)
dict_folder_path = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/dict_inst_2/Named_inst_2/'
new_dict_all_sims = change_name_of_sub_sub_inst(new_dict_all_sims_sub, dict_folder_path)
# save_and_archiv_instances(new_dict_all_sims, dict_of_key_sections) # todo muss umgebaut werden, da die umbenennung in separater Function stattfindet
elapsed = time.time() - t
print(elapsed)

f_4 = OrderedDict()
path = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/folder_for_dict_instances/'
# for name_2 in os.listdir('C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis'):
# modified_all_sims = all_sims
#     if name_2.endswith('.json'):
# dict_to_be_transfered_to_mongo =
mongo_cluster = 'mongodb+srv://HendrikB:tAyJ0AtmHUiqaYNf@cluster0.4a1ca.mongodb.net/myFirstDatabase?retryWrites=true&w=majority'
mongo_db = 'MasterthesisDB'
# mongo_collection = 'instance_JSON_dicts'
mongo_collection_1 = 'unmodified_JSON_dicts'
mongo_collection_2 = 'unmodified_JSON_dicts'
Path_1 = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/folder_for_dict_instances/Named_instances/'
Path_2 = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/Json_Dict_unmodified/'
# transfer_simulation_dicts_to_mongo_db(Path_2, mongo_db, mongo_collection_2, mongo_cluster)


# for root, dirs, files in os.walk(path):
#     for name_2 in files:
#         if name_2.endswith(".json"):
#             with open(root + str('/') + name_2, 'r') as in_file:
#                 # file_name = os.path.normpath(root.split(path)[1] + name_2.split('.out')[0])
#                 # file_name = file_name[1:].replace('\\', '_')
#                 # file_list_1.append([root, name_2, file_name])
#                 test_567 = json.load(in_file)
#                 instance_name = name_2
#                 test_678 = list(root.split('/'))[-1].split('\\')
#                 instance_key = test_678[0]
#                 instance_group = test_678[1]
#                 print('root: ', root)
#                 print('instance_key: ', instance_key)
#                 print('instance_group: ', instance_group)
#                 print('instance_name: ', instance_name)
#                 transfer_data_to_mongo({'instance_name': instance_name,'instance_key': instance_key, 'instance_group': instance_group, 'dict': test_567}, mongo_db, mongo_collection, mongo_cluster)

                # for i_int in test_567['Sims']:
                #     print(test_567['Tuple'])
                #     modified_all_sims_2 = modified_all_sims[i_int]
                #     for i_2_int in test_567['Tuple'][:-1]:
                #         modified_all_sims_2 = modified_all_sims_2[i_2_int]
                #     print(modified_all_sims_2)
                # print(test_567['Tuple'])
        # with open(name_2) as outfile:
        #     f_4.update({name_2.rstrip('.json'): json.load(outfile)})



# cluster = MongoClient('mongodb+srv://HendrikB:tAyJ0AtmHUiqaYNf@cluster0.4a1ca.mongodb.net/myFirstDatabase?retryWrites=true&w=majority')
# db = cluster['MasterthesisDB']
# mongo_db_collection = db['JSONdicts']
# mongo_db_collection.delete_many({})
# mongo_cluster = 'mongodb+srv://HendrikB:tAyJ0AtmHUiqaYNf@cluster0.4a1ca.mongodb.net/myFirstDatabase?retryWrites=true&w=majority'
# mongo_db = 'MasterthesisDB'
# mongo_collection = 'JSON_dicts'

# transfer_data_to_mongo({'_id': 0, 'dict': 'test'}, mongo_db, mongo_collection, mongo_cluster)
#
# for number, json_dict in enumerate(all_sims):
#     transfer_data_to_mongo({'_id': number, 'dict': {json_dict: all_sims[json_dict]}}, mongo_db, mongo_collection, mongo_cluster)




Path_3 = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/Translator_test_modified.xlsx'

df_1 = pd.read_excel(Path_3, sheet_name='For testing on Components 3')

dict_of_excel_translator = df_1.to_dict()
dict_of_excel_translator['Implementation_Check']=dict(zip(
    list(range(len(dict_of_excel_translator['Onto_superclass']))),
    [False]*len(dict_of_excel_translator['Onto_superclass'])))

list_onto_superclass = list(dict_of_excel_translator['Onto_superclass'].values())
list_onto_name = list(dict_of_excel_translator['Onto_name'].values())
list_dict_names = list(dict_of_excel_translator['Dict_name'].values())
list_implementation_check  = list(dict_of_excel_translator['Implementation_Check'].values())
zipped_list = zip(list_onto_superclass,
                  list_onto_name,
                  list_dict_names,
                  list_implementation_check)

with emmo:
    namespace = 'emmo'
    ProgramSettings = types.new_class('ProgramSettings', (emmo.Symbolic,))
    emmo.ProgramSettings.prefLabel = 'ProgramSettings'
    DomainSettings = types.new_class('DomainSettings', (emmo.ProgramSettings,))
    emmo.DomainSettings.prefLabel = 'DomainSettings'
    CfxFunctionSetting = types.new_class('CfxFunctionSetting', (emmo.ProgramSettings,))
    emmo.CfxFunctionSetting.prefLabel = 'CfxFunctionSetting'
    ParallelHostSettings = types.new_class('ParallelHostSettings', (emmo.ProgramSettings,))
    emmo.ParallelHostSettings.prefLabel = 'ParallelHostSettings'
    MaterialSettings = types.new_class('MaterialSettings', (emmo.ProgramSettings,))
    emmo.MaterialSettings.prefLabel = 'MaterialSettings'
    InputFieldSettings = types.new_class('InputFieldSettings', (emmo.ProgramSettings,))
    emmo.InputFieldSettings.prefLabel = 'InputFieldSettings'
    CfxFunctionDataField = types.new_class('CfxFunctionDataField', (emmo.ProgramSettings,))
    emmo.CfxFunctionDataField.prefLabel = 'CfxFunctionDataField'
    MathematicalSimulations = types.new_class('MathematicalSimulations', (emmo.Symbolic,))
    emmo.MathematicalSimulations.prefLabel = 'MathematicalSimulations'

emmo = modify_ontologiy__by_adding_classes(dict_of_excel_translator, emmo)



test_dict_for_translator = copy.deepcopy(all_sims)
search_for_name_option_in_dict(test_dict_for_translator)
list_of_keys_int = list_of_keys_containing_named_sub_sub_sub_instances + list_of_keys_containing_named_sub_sub_instances + list_of_keys_containing_named_sub_instances + list_of_keys_containing_named_instances + list_of_instance_sub_sub_keys + list_of_instance_sub_keys + list_of_instance_keys
test_dict_for_translator_2 = start_rename_dict_chapters_for_translator_3(test_dict_for_translator, list_of_keys_int)


merged_dict_for_translator = create_merged_dict(test_dict_for_translator_2)
flattened_dict = flatten_dict_to_list(merged_dict_for_translator['CFX Command Language for Run'])
df_trans = pd.DataFrame(flattened_dict)

FilePath = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/Translator_test_modified_and_ready_for_testing_2.xlsx'
ExcelWorkbook = load_workbook(FilePath)
writer = pd.ExcelWriter(FilePath, engine='openpyxl')
writer.book = ExcelWorkbook
df_trans.to_excel(writer, header=False, index=False, sheet_name='List_of_new_names_2')
writer.save()
writer.close()

## todo Test area for changing naming for translator dict

test_dict_for_translator_3 = copy.deepcopy(all_sims)
search_for_name_source_in_dict(test_dict_for_translator_3)
# search_for_name_option_in_dict(test_dict_for_translator_3)
# list_of_keys_int_2 = list_of_keys_containing_named_sub_sub_sub_instances + list_of_keys_containing_named_sub_sub_instances + list_of_keys_containing_named_sub_instances + list_of_keys_containing_named_instances + list_of_instance_sub_sub_keys + list_of_instance_sub_keys + list_of_instance_keys
# test_dict_for_translator_4 = start_rename_dict_chapters_for_translator_3(test_dict_for_translator_3, list_of_keys_int_2)
#
#
# merged_dict_for_translator_2 = create_merged_dict(test_dict_for_translator_4)
# flattened_dict_2 = flatten_dict_to_list(merged_dict_for_translator_2['CFX Command Language for Run'])
#
#
# flattened_dict_2 = list(dict.fromkeys(flattened_dict_2))
# df_trans_2 = pd.DataFrame(flattened_dict_2)

# Path_4 = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/Translator_test_modified_and_ready_for_testing.xlsx'
# FilePath = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/Translator_test_modified_and_ready_for_testing_2.xlsx'
# ExcelWorkbook = load_workbook(FilePath)
# writer = pd.ExcelWriter(FilePath, engine='openpyxl')
# writer.book = ExcelWorkbook
# df_trans_2.to_excel(writer, header=False, index=False, sheet_name='List_of_new_names_2')
# writer.save()
# writer.close()
##


flattened_dict = list(dict.fromkeys(flattened_dict))
df_trans = pd.DataFrame(flattened_dict)
# pd.DataFrame(df_trans).to_excel('Translator_test.xlsx', header=False, index=False)

list_dict_names_2 = list(dict_of_excel_translator['Dict_name'].values())

with emmo:
    hasSimulation = types.new_class('hasSimulation', (owlready2.ObjectProperty,))
    hasIndividual = types.new_class('hasIndividual', (owlready2.ObjectProperty,))
    hasDefinedIndividual = types.new_class('hasDefinedIndividual', (owlready2.ObjectProperty,))
    hasSetting = types.new_class('hasSetting', (owlready2.ObjectProperty,))
    hasOption = types.new_class('hasOption', (owlready2.DataProperty,))

    # hasTempDataProp = types.new_class('hasTempDataProp', (owlready2.DataProperty,))

Path_3 = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/dict_inst_2/Named_sub_sub_sub_inst_2'
for root_3, dirs_3, files_3 in os.walk(Path_3):
    for name_3 in files_3:
        with open(root_3 + str('/') + name_3, 'r') as in_file:
            dict_for_emmo_instances = json.load(in_file)
        start_dict_to_emmo_individuals(dict_for_emmo_instances, name_3, dict_of_excel_translator, emmo, test=False)

Path_4 = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/dict_inst_2/Named_sub_sub_inst_2'
for root_4, dirs_4, files_4 in os.walk(Path_4):
    for name_4 in files_4:
        with open(root_4 + str('/') + name_4, 'r') as in_file_2:
            dict_for_emmo_instances_2 = json.load(in_file_2)
        start_dict_to_emmo_individuals(dict_for_emmo_instances_2, name_4, dict_of_excel_translator, emmo, test=True)

Path_5 = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/dict_inst_2/Named_sub_inst_2'
for root_5, dirs_5, files_5 in os.walk(Path_5):
    for name_5 in files_5:
        with open(root_5 + str('/') + name_5, 'r') as in_file_3:
            dict_for_emmo_instances_3 = json.load(in_file_3)
        start_dict_to_emmo_individuals(dict_for_emmo_instances_3, name_5, dict_of_excel_translator, emmo, test=True)

Path_6 = 'C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis/dict_inst_2/Named_inst_2'
for root_6, dirs_6, files_6 in os.walk(Path_6):
    for name_6 in files_6:
        with open(root_6 + str('/') + name_6, 'r') as in_file_4:
            dict_for_emmo_instances_4 = json.load(in_file_4)
        start_dict_to_emmo_individuals(dict_for_emmo_instances_4, name_6, dict_of_excel_translator, emmo, test=True)


emmo.save(os.path.join('C:/Users/hendr/Documents/GitHub/Hendrik_Borgelt_Masterthesis',
                       'emmo_inferred_20220216_trying_classes.owl'))


# dict_of_excel_translator['Onto_name'][48]


elapsed = time.time() - t_2
print(elapsed)

conv_dict = add_own_kpis_and_save_figs(all_sims)

print('Finished')


