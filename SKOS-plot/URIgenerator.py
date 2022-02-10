# -*- coding: utf-8 -*-
"""
Created on Fri Nov 19 12:50:26 2021

@author: Alexander Behr
@affiliation: TU Dortmund Fac. BCI - AG AD

Short description:
The method URI_generation converts entries of preferred label (from column B) 
to concept URIs (into column A). Then it replaces the entries in columns 
children (column E) and related (column F) with their respective URIs.


Description:
First method URI_generation reads in the URI for the vocabulary (in cell B1) 
and the preferred labels. To get a URI-valid name, the space characters are 
then replaced by an underscore and put concatenated with the vocabulary URI 
into the fields of concept URI. While this for loop runs, all preferred labels 
are stored with their respective URIs into a dictionary (URI_dict).
The second loop then iterates through columns 5 and 6 (children and related).
When an entry is found, that is not None and does not contain some of 
(https://, http:// or www.), the entry is split by commas and each of those 
subentries is then replaced by it's URI by the dictionary (URI_dict).

"""



from pathlib import Path
def URI_generation(
        excel_file_path: Path
        ):
    import openpyxl

    #Load excel-file
    wb = openpyxl.load_workbook(excel_file_path)
    
    #open worksheet vocabulary
    ws = wb['vocabulary']
    
    #base URI, defined in header of excel-file
    base_URI = ws["B1"].value
    
    URI_dict = {}
    
    #iterating through the first two columns, starting from the row, where concept definitions start (Row 15)
    # saves dictionary, to map preferred labels to respective URIS
    for row in ws.iter_rows(min_row = 15,max_col = 2):
        #if concept URI is empty (None) and preferred label has a value, derive URI tag from preferred label
        if not row[0].value and row[1].value:
            underscored_label = row[1].value.replace(" ", "_")
            new_URI = base_URI + underscored_label
            row_number = row[0].row
            ws.cell(row=row_number,column=1).value = new_URI
            URI_dict[row[1].value] = new_URI
        
    
    #iterating through columns children and related, replacing entries with concept URIs
    for row in ws.iter_rows(min_row = 16,min_col=5, max_col = 6):
        for entry in row:
            if entry.value and (not ("https://" or "http://" or "www.") in entry.value):
                concept_list = entry.value.split(",")
                URI_list =[]
                for concept in concept_list:
                    # using the concept as key, searching for the URI of the concept
                    # lstrip cuts of all leading whitespace-characters
                    URI_list.append(URI_dict[concept.lstrip()])
                
                # join URI_list to one string
                entry_string = ', '.join(URI_list)
                ws.cell(row=entry.row,column=entry.column).value = entry_string
        
        
    
    # save Excel-file (overwriting)
    wb.save(excel_file_path)
